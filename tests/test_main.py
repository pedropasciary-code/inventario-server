import csv
import io
import json
import re
from datetime import UTC, datetime, timedelta
from types import SimpleNamespace

from openpyxl import load_workbook

from app import models
from app.auth import hash_password, verify_password


AGENT_HEADERS = {"X-Agent-Token": "test-agent-token"}


def extract_csrf_token(response):
    match = re.search(r'name="csrf_token" value="([^"]+)"', response.text)
    assert match, response.text
    return match.group(1)


def login(client):
    response = client.get("/login")
    csrf_token = extract_csrf_token(response)

    return client.post(
        "/login",
        data={
            "username": "admin",
            "password": "strong-password",
            "csrf_token": csrf_token,
        },
        follow_redirects=False,
    )


def test_checkin_creates_and_updates_asset(client, db_session):
    payload = {
        "hostname": "PC-01",
        "serial": "SERIAL-01",
        "mac_address": "AA-BB-CC-00-00-01",
        "ip": "10.0.0.1",
    }

    create_response = client.post("/checkin", json=payload, headers=AGENT_HEADERS)
    assert create_response.status_code == 200
    assert create_response.json()["hostname"] == "PC-01"

    update_response = client.post(
        "/checkin",
        json={**payload, "hostname": "PC-01-RENAMED", "ip": "10.0.0.2"},
        headers=AGENT_HEADERS,
    )
    assert update_response.status_code == 200
    assert update_response.json()["id"] == create_response.json()["id"]
    assert update_response.json()["hostname"] == "PC-01-RENAMED"
    assert db_session.query(models.Asset).count() == 1
    checkins = db_session.query(models.AssetCheckin).order_by(models.AssetCheckin.id.asc()).all()
    assert [checkin.event_type for checkin in checkins] == ["created", "updated"]
    assert checkins[-1].hostname == "PC-01-RENAMED"


def test_checkin_normalizes_serial_and_mac_identity(client, db_session):
    create_response = client.post(
        "/checkin",
        json={
            "hostname": "PC-01",
            "serial": " serial-01 ",
            "mac_address": "aa:bb:cc:00:00:01",
        },
        headers=AGENT_HEADERS,
    )
    assert create_response.status_code == 200

    update_response = client.post(
        "/checkin",
        json={
            "hostname": "PC-01-UPDATED",
            "serial": "SERIAL-01",
            "mac_address": "AA-BB-CC-00-00-01",
        },
        headers=AGENT_HEADERS,
    )

    assert update_response.status_code == 200
    assert update_response.json()["id"] == create_response.json()["id"]
    asset = db_session.query(models.Asset).one()
    assert asset.serial == "SERIAL-01"
    assert asset.mac_address == "AA-BB-CC-00-00-01"


def test_checkin_rejects_invalid_agent_token(client):
    response = client.post(
        "/checkin",
        json={"hostname": "PC-01"},
        headers={"X-Agent-Token": "wrong-token"},
    )

    assert response.status_code == 401


def test_checkin_rejects_conflicting_identity(client):
    first = {
        "hostname": "PC-01",
        "serial": "SERIAL-01",
        "mac_address": "AA-BB-CC-00-00-01",
    }
    second = {
        "hostname": "PC-02",
        "serial": "SERIAL-02",
        "mac_address": "AA-BB-CC-00-00-02",
    }

    assert client.post("/checkin", json=first, headers=AGENT_HEADERS).status_code == 200
    assert client.post("/checkin", json=second, headers=AGENT_HEADERS).status_code == 200

    conflict_response = client.post(
        "/checkin",
        json={
            "hostname": "PC-CONFLICT",
            "serial": "SERIAL-01",
            "mac_address": "AA-BB-CC-00-00-02",
        },
        headers=AGENT_HEADERS,
    )

    assert conflict_response.status_code == 409


def test_checkin_conflict_records_audit_event(client, db_session):
    first = {
        "hostname": "PC-01",
        "serial": "SERIAL-01",
        "mac_address": "AA-BB-CC-00-00-01",
    }
    second = {
        "hostname": "PC-02",
        "serial": "SERIAL-02",
        "mac_address": "AA-BB-CC-00-00-02",
    }

    assert client.post("/checkin", json=first, headers=AGENT_HEADERS).status_code == 200
    assert client.post("/checkin", json=second, headers=AGENT_HEADERS).status_code == 200

    response = client.post(
        "/checkin",
        json={
            "hostname": "PC-CONFLICT",
            "serial": "SERIAL-01",
            "mac_address": "AA-BB-CC-00-00-02",
        },
        headers=AGENT_HEADERS,
    )

    assert response.status_code == 409
    event = db_session.query(models.AuditEvent).filter_by(event_type="checkin_rejected").one()
    details = json.loads(event.details_json)
    assert details["reason"] == "identity_conflict"
    assert details["serial"] == "SERIAL-01"


def test_checkin_rate_limit_blocks_flood(client):
    payload = {"hostname": "PC-FLOOD", "mac_address": "AA-BB-CC-00-FF-01"}

    from app.rate_limiting import CHECKIN_RATE_LIMIT_MAX

    for i in range(CHECKIN_RATE_LIMIT_MAX):
        response = client.post("/checkin", json=payload, headers=AGENT_HEADERS)
        assert response.status_code == 200

    blocked = client.post(
        "/checkin",
        json=payload,
        headers=AGENT_HEADERS,
    )
    assert blocked.status_code == 429


def test_checkin_rate_limit_allows_distinct_assets_behind_same_ip(client):
    from app.rate_limiting import CHECKIN_RATE_LIMIT_MAX, checkin_attempts

    for i in range(CHECKIN_RATE_LIMIT_MAX + 1):
        response = client.post(
            "/checkin",
            json={"hostname": f"PC-NAT-{i}", "mac_address": f"AA-BB-CC-00-FE-{i:02d}"},
            headers=AGENT_HEADERS,
        )
        assert response.status_code == 200

    assert len(checkin_attempts) == CHECKIN_RATE_LIMIT_MAX + 1


def test_login_requires_valid_csrf_token(client, admin_user):
    response = client.get("/login")
    assert response.status_code == 200

    invalid_response = client.post(
        "/login",
        data={
            "username": "admin",
            "password": "strong-password",
            "csrf_token": "invalid",
        },
    )

    assert invalid_response.status_code == 403


def test_login_rate_limit_blocks_repeated_failures(client, admin_user):
    csrf_token = extract_csrf_token(client.get("/login"))

    for _ in range(5):
        response = client.post(
            "/login",
            data={
                "username": "admin",
                "password": "wrong-password",
                "csrf_token": csrf_token,
            },
        )
        assert response.status_code == 401

    blocked_response = client.post(
        "/login",
        data={
            "username": "admin",
            "password": "wrong-password",
            "csrf_token": csrf_token,
        },
    )

    assert blocked_response.status_code == 429


def test_login_success_and_failure_record_audit_events(client, db_session, admin_user):
    csrf_token = extract_csrf_token(client.get("/login"))
    failed_response = client.post(
        "/login",
        data={
            "username": "admin",
            "password": "wrong-password",
            "csrf_token": csrf_token,
        },
        follow_redirects=False,
    )
    assert failed_response.status_code == 401

    success_response = login(client)
    assert success_response.status_code == 303

    events = db_session.query(models.AuditEvent).order_by(models.AuditEvent.id.asc()).all()
    assert [event.event_type for event in events] == ["login_failed", "login_success"]
    assert events[0].username == "admin"
    assert events[1].username == "admin"


def test_login_rehashes_legacy_pbkdf2_password(client, db_session):
    import hashlib

    password = "legacy-password"
    salt = "legacy-salt"
    iterations = 100_000
    password_hash = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt.encode("utf-8"),
        iterations,
    ).hex()
    user = models.User(
        username="legacy",
        password_hash=f"{iterations}${salt}${password_hash}",
        is_active=True,
        is_admin=False,
    )
    db_session.add(user)
    db_session.commit()

    csrf_token = extract_csrf_token(client.get("/login"))
    response = client.post(
        "/login",
        data={"username": "legacy", "password": password, "csrf_token": csrf_token},
        follow_redirects=False,
    )

    assert response.status_code == 303
    db_session.refresh(user)
    assert user.password_hash.startswith("$argon2")
    assert verify_password(password, user.password_hash) is True


def test_dashboard_paginates_and_shows_real_status(client, db_session, admin_user):
    now = datetime.now(UTC)

    for index in range(30):
        if index % 3 == 0:
            last_seen = now
        elif index % 3 == 1:
            last_seen = now - timedelta(days=2)
        else:
            last_seen = now - timedelta(days=10)

        db_session.add(
            models.Asset(
                hostname=f"PC-{index:02d}",
                serial=f"SERIAL-{index:02d}",
                mac_address=f"AA-BB-CC-00-00-{index:02d}",
                ultima_comunicacao=last_seen,
            )
        )

    db_session.commit()

    login_response = login(client)
    assert login_response.status_code == 303

    response = client.get("/dashboard?page=2&per_page=10")

    assert response.status_code == 200
    assert "Exibindo 11-20 de 30 ativos" in response.text
    assert "Página 2 de 3" in response.text
    assert "Comunicando" in response.text
    assert "Atrasado" in response.text
    assert "Inativo" in response.text


def test_asset_detail_shows_network_interfaces(client, db_session, admin_user):
    interfaces = [
        {
            "name": "Ethernet",
            "mac_address": "AA-BB-CC-00-00-01",
            "ip_addresses": ["192.168.0.10"],
            "is_virtual": False,
        },
        {
            "name": "VPN Adapter",
            "mac_address": "AA-BB-CC-00-00-02",
            "ip_addresses": ["10.8.0.2"],
            "is_virtual": True,
        },
    ]
    asset = models.Asset(
        hostname="PC-DETAIL",
        serial="SERIAL-DETAIL",
        mac_address="AA-BB-CC-00-00-01",
        network_interfaces=json.dumps(interfaces),
        ultima_comunicacao=datetime.now(UTC),
    )
    db_session.add(asset)
    db_session.flush()
    db_session.add(
        models.AssetCheckin(
            asset_id=asset.id,
            event_type="created",
            hostname="PC-DETAIL",
            ip="192.168.0.10",
            mac_address="AA-BB-CC-00-00-01",
            agent_version="1.0.0",
            payload_json="{}",
            created_at=datetime.now(UTC),
        )
    )
    db_session.commit()
    db_session.refresh(asset)

    assert login(client).status_code == 303

    response = client.get(f"/assets/{asset.id}")

    assert response.status_code == 200
    assert "Interfaces de Rede" in response.text
    assert "Ethernet" in response.text
    assert "VPN Adapter" in response.text
    assert "Física" in response.text
    assert "Virtual" in response.text
    assert "Histórico de Check-ins" in response.text
    assert "PC-DETAIL" in response.text
    assert "1.0.0" in response.text


def test_dashboard_filters_status_and_sorts(client, db_session, admin_user):
    now = datetime.now(UTC)
    db_session.add_all(
        [
            models.Asset(
                hostname="PC-ONLINE",
                usuario="ana",
                serial="SERIAL-ONLINE",
                mac_address="AA-BB-CC-00-10-01",
                ultima_comunicacao=now,
            ),
            models.Asset(
                hostname="PC-STALE",
                usuario="bruno",
                serial="SERIAL-STALE",
                mac_address="AA-BB-CC-00-10-02",
                ultima_comunicacao=now - timedelta(days=2),
            ),
            models.Asset(
                hostname="PC-INACTIVE",
                usuario="carla",
                serial="SERIAL-INACTIVE",
                mac_address="AA-BB-CC-00-10-03",
                ultima_comunicacao=now - timedelta(days=10),
            ),
        ]
    )
    db_session.commit()

    assert login(client).status_code == 303

    response = client.get("/dashboard?status=stale&sort=usuario&direction=desc&per_page=10")

    assert response.status_code == 200
    assert "PC-STALE" in response.text
    assert "PC-ONLINE" not in response.text
    assert "PC-INACTIVE" not in response.text
    assert "Atrasado" in response.text
    assert "10 por página" in response.text


def test_dashboard_warns_when_export_will_be_truncated(client, db_session, admin_user, monkeypatch):
    from app.routers import dashboard as dashboard_router

    monkeypatch.setattr(dashboard_router, "EXPORT_MAX_ROWS", 1)
    now = datetime.now(UTC)
    db_session.add_all(
        [
            models.Asset(
                hostname="PC-EXPORT-1",
                serial="SERIAL-EXPORT-1",
                mac_address="AA-BB-CC-00-11-01",
                ultima_comunicacao=now,
            ),
            models.Asset(
                hostname="PC-EXPORT-2",
                serial="SERIAL-EXPORT-2",
                mac_address="AA-BB-CC-00-11-02",
                ultima_comunicacao=now,
            ),
        ]
    )
    db_session.commit()

    assert login(client).status_code == 303
    response = client.get("/dashboard")

    assert response.status_code == 200
    assert "A exportacao deste filtro sera limitada" in response.text
    assert "primeiros 1 registros" in response.text


def test_csv_export_includes_status_and_formatted_dates(client, db_session, admin_user):
    now = datetime.now(UTC)
    db_session.add_all(
        [
            models.Asset(
                hostname="PC-ONLINE",
                serial="SERIAL-ONLINE",
                mac_address="AA-BB-CC-00-20-01",
                ultima_comunicacao=now,
            ),
            models.Asset(
                hostname="PC-INACTIVE",
                serial="SERIAL-INACTIVE",
                mac_address="AA-BB-CC-00-20-02",
                ultima_comunicacao=now - timedelta(days=10),
            ),
        ]
    )
    db_session.commit()

    assert login(client).status_code == 303

    response = client.get("/export/csv?status=inactive")

    assert response.status_code == 200
    rows = list(csv.DictReader(io.StringIO(response.text)))
    assert len(rows) == 1
    assert rows[0]["Hostname"] == "PC-INACTIVE"
    assert rows[0]["Status"] == "Inativo"
    assert "/" in rows[0]["Ultima Comunicacao"]
    assert response.headers["X-Export-Row-Limit"] == "10000"
    assert response.headers["X-Export-Truncated"] == "false"


def test_exports_and_logout_record_audit_events(client, db_session, admin_user):
    db_session.add(
        models.Asset(
            hostname="PC-AUDIT",
            serial="SERIAL-AUDIT",
            mac_address="AA-BB-CC-00-30-01",
            ultima_comunicacao=datetime.now(UTC),
        )
    )
    db_session.commit()

    assert login(client).status_code == 303
    assert client.get("/export/csv?status=online&sort=hostname&direction=asc").status_code == 200
    assert client.get("/export/xlsx?status=all&sort=hostname&direction=asc").status_code == 200

    dashboard_response = client.get("/dashboard")
    csrf_token = extract_csrf_token(dashboard_response)
    logout_response = client.post(
        "/logout",
        data={"csrf_token": csrf_token},
        follow_redirects=False,
    )
    assert logout_response.status_code == 303

    events = db_session.query(models.AuditEvent).order_by(models.AuditEvent.id.asc()).all()
    event_types = [event.event_type for event in events]
    assert event_types == ["login_success", "export_csv", "export_xlsx", "logout"]
    assert all(event.username == "admin" for event in events)
    export_details = json.loads(events[1].details_json)
    assert export_details["status"] == "online"


def test_audit_page_requires_login(client):
    response = client.get("/audit", follow_redirects=False)

    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_audit_page_requires_admin(client, db_session):
    user = models.User(
        username="operator",
        password_hash=hash_password("operator-password"),
        is_active=True,
        is_admin=False,
    )
    db_session.add(user)
    db_session.commit()

    response = client.get("/login")
    csrf_token = extract_csrf_token(response)
    login_response = client.post(
        "/login",
        data={
            "username": "operator",
            "password": "operator-password",
            "csrf_token": csrf_token,
        },
        follow_redirects=False,
    )
    assert login_response.status_code == 303

    response = client.get("/audit")

    assert response.status_code == 403


def test_audit_page_filters_events(client, db_session, admin_user):
    db_session.add_all(
        [
            models.AuditEvent(
                event_type="checkin_rejected",
                username=None,
                ip_address="127.0.0.1",
                details_json=json.dumps({"reason": "identity_conflict"}),
                created_at=datetime.now(UTC),
            ),
            models.AuditEvent(
                event_type="export_csv",
                username="admin",
                ip_address="127.0.0.1",
                details_json=json.dumps({"status": "all"}),
                created_at=datetime.now(UTC),
            ),
        ]
    )
    db_session.commit()

    assert login(client).status_code == 303

    response = client.get("/audit?event_type=checkin_rejected&per_page=10")

    assert response.status_code == 200
    assert "Check-in rejeitado" in response.text
    assert "identity_conflict" in response.text
    assert '"status": "all"' not in response.text


def test_audit_csv_export_respects_filters(client, db_session, admin_user):
    db_session.add_all(
        [
            models.AuditEvent(
                event_type="checkin_rejected",
                username=None,
                ip_address="127.0.0.1",
                details_json=json.dumps({"reason": "missing_identity"}),
                created_at=datetime.now(UTC),
            ),
            models.AuditEvent(
                event_type="login_failed",
                username="admin",
                ip_address="127.0.0.1",
                details_json=json.dumps({"reason": "invalid_credentials"}),
                created_at=datetime.now(UTC),
            ),
        ]
    )
    db_session.commit()

    assert login(client).status_code == 303

    response = client.get("/audit/export/csv?event_type=login_failed")

    assert response.status_code == 200
    rows = list(csv.DictReader(io.StringIO(response.text)))
    assert len(rows) == 1
    assert rows[0]["Tipo"] == "login_failed"
    assert rows[0]["Usuario"] == "admin"


def test_audit_xlsx_export_respects_filters(client, db_session, admin_user):
    db_session.add_all(
        [
            models.AuditEvent(
                event_type="export_csv",
                username="admin",
                ip_address="127.0.0.1",
                details_json=json.dumps({"status": "all"}),
                created_at=datetime.now(UTC),
            ),
            models.AuditEvent(
                event_type="login_failed",
                username="operator",
                ip_address="127.0.0.2",
                details_json=json.dumps({"reason": "invalid_credentials"}),
                created_at=datetime.now(UTC),
            ),
        ]
    )
    db_session.commit()

    assert login(client).status_code == 303

    response = client.get("/audit/export/xlsx?event_type=export_csv")

    assert response.status_code == 200
    assert response.headers["X-Export-Row-Limit"] == "10000"
    assert response.headers["X-Export-Truncated"] == "false"
    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("Data", "Tipo", "Usuario", "IP", "Detalhes")
    assert len(rows) == 2
    assert rows[1][1] == "export_csv"
    assert rows[1][2] == "admin"


def test_users_page_requires_login(client):
    response = client.get("/users", follow_redirects=False)

    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_users_page_requires_admin(client, db_session):
    user = models.User(
        username="operator",
        password_hash=hash_password("operator-password"),
        is_active=True,
        is_admin=False,
    )
    db_session.add(user)
    db_session.commit()

    response = client.get("/login")
    csrf_token = extract_csrf_token(response)
    login_response = client.post(
        "/login",
        data={
            "username": "operator",
            "password": "operator-password",
            "csrf_token": csrf_token,
        },
        follow_redirects=False,
    )
    assert login_response.status_code == 303

    response = client.get("/users")

    assert response.status_code == 403


def test_users_page_creates_user_and_records_audit(client, db_session, admin_user):
    assert login(client).status_code == 303
    csrf_token = extract_csrf_token(client.get("/users"))

    response = client.post(
        "/users",
        data={
            "username": "operator",
            "password": "operator-password",
            "password_confirmation": "operator-password",
            "csrf_token": csrf_token,
        },
    )

    assert response.status_code == 200
    assert "Usuario criado com sucesso." in response.text
    user = db_session.query(models.User).filter_by(username="operator").one()
    assert user.is_active is True
    event = db_session.query(models.AuditEvent).filter_by(event_type="user_created").one()
    details = json.loads(event.details_json)
    assert event.username == "admin"
    assert details["target_username"] == "operator"


def test_users_page_rejects_duplicate_user(client, db_session, admin_user):
    assert login(client).status_code == 303
    csrf_token = extract_csrf_token(client.get("/users"))

    response = client.post(
        "/users",
        data={
            "username": "admin",
            "password": "another-password",
            "password_confirmation": "another-password",
            "csrf_token": csrf_token,
        },
    )

    assert response.status_code == 200
    assert "Usuario ja existe." in response.text
    assert db_session.query(models.User).count() == 1


def test_users_page_prevents_disabling_current_user(client, db_session, admin_user):
    assert login(client).status_code == 303
    csrf_token = extract_csrf_token(client.get("/users"))

    response = client.post(
        f"/users/{admin_user.id}/toggle",
        data={"csrf_token": csrf_token},
    )

    assert response.status_code == 200
    assert "Voce nao pode desativar o proprio usuario logado." in response.text
    db_session.refresh(admin_user)
    assert admin_user.is_active is True


def test_users_page_toggles_user_and_changes_password(client, db_session, admin_user):
    operator = models.User(
        username="operator",
        password_hash="old-hash",
        is_active=True,
    )
    db_session.add(operator)
    db_session.commit()
    db_session.refresh(operator)

    assert login(client).status_code == 303
    csrf_token = extract_csrf_token(client.get("/users"))

    disable_response = client.post(
        f"/users/{operator.id}/toggle",
        data={"csrf_token": csrf_token},
    )
    assert disable_response.status_code == 200
    db_session.refresh(operator)
    assert operator.is_active is False

    password_response = client.post(
        f"/users/{operator.id}/password",
        data={
            "password": "new-operator-password",
            "password_confirmation": "new-operator-password",
            "csrf_token": csrf_token,
        },
    )
    assert password_response.status_code == 200
    db_session.refresh(operator)
    assert operator.password_hash != "old-hash"

    event_types = [
        event.event_type
        for event in db_session.query(models.AuditEvent).order_by(models.AuditEvent.id.asc()).all()
    ]
    assert event_types == ["login_success", "user_disabled", "password_changed"]


def test_users_page_promotes_and_demotes_admin(client, db_session, admin_user):
    operator = models.User(
        username="operator",
        password_hash="old-hash",
        is_active=True,
        is_admin=False,
    )
    db_session.add(operator)
    db_session.commit()
    db_session.refresh(operator)

    assert login(client).status_code == 303
    csrf_token = extract_csrf_token(client.get("/users"))

    promote_response = client.post(
        f"/users/{operator.id}/admin",
        data={"csrf_token": csrf_token},
    )
    assert promote_response.status_code == 200
    db_session.refresh(operator)
    assert operator.is_admin is True

    demote_response = client.post(
        f"/users/{operator.id}/admin",
        data={"csrf_token": csrf_token},
    )
    assert demote_response.status_code == 200
    db_session.refresh(operator)
    assert operator.is_admin is False

    event_types = [
        event.event_type
        for event in db_session.query(models.AuditEvent).order_by(models.AuditEvent.id.asc()).all()
    ]
    assert event_types == ["login_success", "user_promoted", "user_demoted"]


def test_users_page_prevents_demoting_current_admin(client, db_session, admin_user):
    assert login(client).status_code == 303
    csrf_token = extract_csrf_token(client.get("/users"))

    response = client.post(
        f"/users/{admin_user.id}/admin",
        data={"csrf_token": csrf_token},
    )

    assert response.status_code == 200
    assert "Voce nao pode remover o proprio acesso admin." in response.text
    db_session.refresh(admin_user)
    assert admin_user.is_admin is True


def test_apply_asset_payload_preserves_existing_fields_when_none(client, db_session):
    # First checkin sets ultimo_boot
    payload = {
        "hostname": "PC-PRESERVE",
        "serial": "SERIAL-PRESERVE",
        "mac_address": "AA-BB-CC-00-AA-01",
        "ultimo_boot": "2026-01-15T08:00:00Z",
    }
    response = client.post("/checkin", json=payload, headers=AGENT_HEADERS)
    assert response.status_code == 200

    asset = db_session.query(models.Asset).one()
    assert asset.ultimo_boot is not None

    # Second checkin omits ultimo_boot — existing value must be preserved
    response = client.post(
        "/checkin",
        json={"hostname": "PC-PRESERVE", "serial": "SERIAL-PRESERVE", "mac_address": "AA-BB-CC-00-AA-01"},
        headers=AGENT_HEADERS,
    )
    assert response.status_code == 200
    db_session.refresh(asset)
    assert asset.ultimo_boot is not None, "ultimo_boot was erased by a payload without it"


def test_checkin_allows_explicit_null_to_clear_field(client, db_session):
    payload = {
        "hostname": "PC-CLEAR",
        "serial": "SERIAL-CLEAR",
        "mac_address": "AA-BB-CC-00-AA-02",
        "usuario": "alice",
    }
    response = client.post("/checkin", json=payload, headers=AGENT_HEADERS)
    assert response.status_code == 200

    response = client.post(
        "/checkin",
        json={
            "hostname": "PC-CLEAR",
            "serial": "SERIAL-CLEAR",
            "mac_address": "AA-BB-CC-00-AA-02",
            "usuario": None,
        },
        headers=AGENT_HEADERS,
    )

    assert response.status_code == 200
    asset = db_session.query(models.Asset).one()
    assert asset.usuario is None


def test_checkin_omitted_identity_fields_are_preserved(client, db_session):
    payload = {
        "hostname": "PC-IDENTITY",
        "serial": "SERIAL-IDENTITY",
        "mac_address": "AA-BB-CC-00-AA-03",
    }
    assert client.post("/checkin", json=payload, headers=AGENT_HEADERS).status_code == 200

    response = client.post(
        "/checkin",
        json={"hostname": "PC-IDENTITY", "ip": "10.0.0.10"},
        headers=AGENT_HEADERS,
    )

    assert response.status_code == 200
    asset = db_session.query(models.Asset).one()
    assert asset.serial == "SERIAL-IDENTITY"
    assert asset.mac_address == "AA-BB-CC-00-AA-03"
    assert asset.ip == "10.0.0.10"


def test_verify_password_accepts_legacy_pbkdf2_hash():
    import hashlib

    password = "legacy-password"
    salt = "legacy-salt"
    iterations = 100_000
    password_hash = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt.encode("utf-8"),
        iterations,
    ).hex()

    assert verify_password(password, f"{iterations}${salt}${password_hash}") is True
    assert verify_password("wrong-password", f"{iterations}${salt}${password_hash}") is False


def test_parse_non_negative_int_env_rejects_invalid_values(monkeypatch):
    from app.config import parse_non_negative_int_env, parse_positive_int_env

    monkeypatch.setenv("CHECKIN_RETENTION_DAYS", "abc")
    try:
        parse_non_negative_int_env("CHECKIN_RETENTION_DAYS", 90)
    except ValueError as error:
        assert "deve ser um numero inteiro" in str(error)
    else:
        raise AssertionError("invalid retention value was accepted")

    monkeypatch.setenv("CHECKIN_RETENTION_DAYS", "-1")
    try:
        parse_non_negative_int_env("CHECKIN_RETENTION_DAYS", 90)
    except ValueError as error:
        assert "nao pode ser negativo" in str(error)
    else:
        raise AssertionError("negative retention value was accepted")

    monkeypatch.setenv("CHECKIN_RATE_LIMIT_MAX", "0")
    try:
        parse_positive_int_env("CHECKIN_RATE_LIMIT_MAX", 30)
    except ValueError as error:
        assert "deve ser maior que zero" in str(error)
    else:
        raise AssertionError("zero rate limit was accepted")


def test_trusted_proxy_header_used_for_rate_limiting(client):
    # Exhaust rate limit from one forged IP via X-Forwarded-For (must NOT bypass limit
    # since ProxyHeadersMiddleware is not active in tests — raw client IP is used)
    from app.rate_limiting import CHECKIN_RATE_LIMIT_MAX, checkin_attempts

    payload = {"hostname": "PC-PROXY", "mac_address": "AA-BB-CC-01-00-00"}
    headers = {**AGENT_HEADERS, "X-Forwarded-For": "198.51.100.10"}

    for i in range(CHECKIN_RATE_LIMIT_MAX):
        r = client.post(
            "/checkin",
            json=payload,
            headers=headers,
        )
        assert r.status_code == 200

    blocked = client.post(
        "/checkin",
        json=payload,
        headers={**AGENT_HEADERS, "X-Forwarded-For": "203.0.113.99"},
    )
    assert blocked.status_code == 429
    # Confirmed: untrusted X-Forwarded-For does not create separate client buckets.
    assert len(checkin_attempts) == 1


def test_record_audit_event_does_not_propagate_db_error(client, db_session, admin_user):
    # Drop the audit_events table to force a DB error on record_audit_event
    from sqlalchemy import text
    db_session.execute(text("DROP TABLE audit_events"))
    db_session.commit()

    # Login should still succeed even though the audit commit will fail
    response = client.get("/login")
    csrf_token = extract_csrf_token(response)
    login_response = client.post(
        "/login",
        data={"username": "admin", "password": "strong-password", "csrf_token": csrf_token},
        follow_redirects=False,
    )
    assert login_response.status_code == 303


def test_record_audit_event_does_not_commit_caller_session(db_session):
    from app.services.audit import record_audit_event

    db_session.add(
        models.Asset(
            hostname="PC-PENDING",
            serial="SERIAL-PENDING",
            mac_address="AA-BB-CC-00-CC-01",
            ultima_comunicacao=datetime.now(UTC),
        )
    )
    request = SimpleNamespace(client=SimpleNamespace(host="127.0.0.1"))

    record_audit_event(db_session, "login_success", request, username="admin")
    db_session.rollback()

    assert db_session.query(models.Asset).filter_by(hostname="PC-PENDING").count() == 0
    event = db_session.query(models.AuditEvent).filter_by(event_type="login_success").one()
    assert event.username == "admin"


def test_checkin_merge_resolves_integrity_conflict(db_session):
    # The merge path in commit_asset_checkin handles IntegrityError that arises
    # during concurrent inserts (race condition). We trigger it at the service
    # layer: one asset exists (only serial), we try to flush a new asset that
    # carries the same serial, causing a unique constraint violation.
    from app.services.asset import commit_asset_checkin

    existing = models.Asset(serial="SERIAL-MERGE-01", hostname="PC-EXISTING")
    db_session.add(existing)
    db_session.commit()

    # Simulate a new asset arriving (find returned None due to race), carrying
    # the same serial that was just inserted by a concurrent request.
    duplicate = models.Asset(serial="SERIAL-MERGE-01", hostname="PC-DUPLICATE", mac_address="AA-BB-CC-00-BB-01")
    db_session.add(duplicate)

    asset_data = {"serial": "SERIAL-MERGE-01", "hostname": "PC-DUPLICATE", "mac_address": "AA-BB-CC-00-BB-01"}
    merged = commit_asset_checkin(duplicate, asset_data, db_session, "created")

    checkins = db_session.query(models.AssetCheckin).order_by(models.AssetCheckin.id).all()
    event_types = [c.event_type for c in checkins]
    assert "merged" in event_types
    assert merged.serial == "SERIAL-MERGE-01"
